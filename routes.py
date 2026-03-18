"""
routes.py - All Flask routes for DigiLib
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import login_user, logout_user, login_required, current_user
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from models import db, Admin, Student, Seat, Shift, Booking, Payment, FeeReminder, NotificationLog
import random, string, io

main = Blueprint('main', __name__)

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def gen_receipt():
    s = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    return f"DL-{datetime.now().year}-{s}"

def calc_discount(original, dtype, dval):
    if dtype == 'percent': disc = round(original * float(dval) / 100, 2)
    elif dtype == 'flat':  disc = min(float(dval), original)
    else:                  disc = 0.0
    return disc, round(original - disc, 2)

def get_stats():
    total    = 92
    students = Student.query.count()
    occupied = Booking.query.filter_by(is_active=True).count()
    revenue  = db.session.query(db.func.sum(Payment.final_amount)).filter_by(status='Paid').scalar() or 0
    today    = date.today()
    daily    = db.session.query(db.func.sum(Payment.final_amount)).filter(
        Payment.status == 'Paid', db.func.date(Payment.payment_date) == today
    ).scalar() or 0
    pending_r = FeeReminder.query.filter_by(status='Pending').count()
    overdue_r = FeeReminder.query.filter(FeeReminder.status=='Pending', FeeReminder.due_date < today).count()

    shifts = Shift.query.all()
    shift_stats = []
    for s in shifts:
        c = Booking.query.filter_by(shift_id=s.id, is_active=True).count()
        shift_stats.append({'name': s.name, 'time': f"{s.start_time} – {s.end_time}",
                            'occupied': c, 'available': total - c, 'percent': round(c / total * 100)})
    return dict(total_seats=total, occupied=occupied, available=total*3-occupied,
                total_students=students, total_revenue=revenue, daily_revenue=daily,
                shift_stats=shift_stats, pending_reminders=pending_r, overdue_reminders=overdue_r)

def ensure_reminder(student):
    months = student.months_active()
    if months < 1: return
    today = date.today()
    reg_day = student.registered_at.day
    try:    due = today.replace(day=reg_day)
    except: due = today.replace(day=28)
    if not FeeReminder.query.filter_by(student_id=student.id, month_number=months).first():
        r = FeeReminder(student_id=student.id, month_number=months, due_date=due,
                        renewal_amount=student.base_fee(), final_renewal_amount=student.base_fee(),
                        status='Overdue' if today > due else 'Pending')
        db.session.add(r)
        db.session.commit()

# ─────────────────────────────────────────────
# PUBLIC ROUTES
# ─────────────────────────────────────────────

@main.route('/')
def index():
    for s in Student.query.all(): ensure_reminder(s)
    return render_template('index.html', stats=get_stats())

@main.route('/seats')
def seats():
    shift_id = request.args.get('shift_id', 1, type=int)
    shifts   = Shift.query.all()
    selected = Shift.query.get_or_404(shift_id)
    seats_data = []
    for n in range(1, 93):
        seat = Seat.query.filter_by(seat_number=n).first()
        status, info = 'available', None
        if seat:
            b = Booking.query.filter_by(seat_id=seat.id, shift_id=shift_id, is_active=True).first()
            if b:
                status = 'occupied'
                p = Payment.query.filter_by(student_id=b.student_id).first()
                info = dict(name=b.student.name, mobile=b.student.mobile, shift=b.shift.name,
                            payment=p.status if p else 'Unpaid', receipt=p.receipt_number if p else 'N/A',
                            mode=p.payment_mode if p else 'N/A')
        seats_data.append(dict(number=n, status=status, student=info))
    return render_template('seats.html', seats=seats_data, shifts=shifts, selected_shift=selected)

@main.route('/register', methods=['GET', 'POST'])
def register():
    shifts = Shift.query.all()
    if request.method == 'POST':
        name            = request.form.get('name', '').strip()
        mobile          = request.form.get('mobile', '').strip()
        dob_str         = request.form.get('dob', '')
        num_shifts      = int(request.form.get('num_shifts', 1))
        shift_ids       = request.form.getlist('shift_ids')
        seat_numbers    = request.form.getlist('seat_numbers')
        discount_type   = request.form.get('discount_type', 'none')
        discount_value  = float(request.form.get('discount_value', 0) or 0)
        discount_reason = request.form.get('discount_reason', '').strip()
        payment_mode    = request.form.get('payment_mode', 'Offline')
        transaction_id  = request.form.get('transaction_id', '').strip()
        mark_paid       = request.form.get('mark_paid') == '1'

        # Validations
        errors = []
        if not all([name, mobile, dob_str]):       errors.append('All fields are required.')
        if len(mobile) != 10 or not mobile.isdigit(): errors.append('Mobile must be 10 digits.')
        if Student.query.filter_by(mobile=mobile).first(): errors.append('Mobile already registered.')
        if len(shift_ids) != num_shifts:           errors.append(f'Select exactly {num_shifts} shift(s).')
        if len(seat_numbers) != num_shifts:        errors.append(f'Select a seat for each shift.')
        if payment_mode == 'Online' and not transaction_id: errors.append('Transaction ID required for online payment.')
        for sn in seat_numbers:
            if not sn.isdigit() or not (1 <= int(sn) <= 92): errors.append(f'Seat {sn} invalid (1–92).')

        if errors:
            for e in errors: flash(e, 'danger')
            return render_template('register.html', shifts=shifts)

        try:    dob = datetime.strptime(dob_str, '%Y-%m-%d').date()
        except: flash('Invalid date of birth.', 'danger'); return render_template('register.html', shifts=shifts)

        # Check seat availability
        for sid, sn in zip(shift_ids, seat_numbers):
            seat = Seat.query.filter_by(seat_number=int(sn)).first()
            if seat and Booking.query.filter_by(seat_id=seat.id, shift_id=int(sid), is_active=True).first():
                sh = Shift.query.get(int(sid))
                flash(f'Seat {sn} already booked for {sh.name}.', 'danger')
                return render_template('register.html', shifts=shifts)

        # Create student
        student = Student(name=name, mobile=mobile, dob=dob, num_shifts=num_shifts)
        db.session.add(student); db.session.flush()

        for sid, sn in zip(shift_ids, seat_numbers):
            seat = Seat.query.filter_by(seat_number=int(sn)).first()
            if not seat:
                seat = Seat(seat_number=int(sn)); db.session.add(seat); db.session.flush()
            db.session.add(Booking(student_id=student.id, seat_id=seat.id, shift_id=int(sid)))

        # Fee calculation
        original = student.base_fee()
        disc_amt, final = calc_discount(original, discount_type, discount_value)
        next_due = (datetime.utcnow() + relativedelta(months=1)).date()

        payment = Payment(
            student_id=student.id, receipt_number=gen_receipt(),
            original_amount=original, discount_type=discount_type,
            discount_value=discount_value, discount_reason=discount_reason,
            discount_amount=disc_amt, final_amount=final,
            status='Paid' if mark_paid else 'Unpaid',
            payment_mode=payment_mode,
            transaction_id=transaction_id or None,
            payment_date=datetime.utcnow() if mark_paid else None,
            next_due_date=next_due, renewal_month=1,
        )
        db.session.add(payment); db.session.commit()

        # Send WhatsApp/SMS notifications
        try:
            from notifications import notify_registration, notify_payment_confirmed
            bookings = Booking.query.filter_by(student_id=student.id, is_active=True).all()
            notify_registration(student, payment, bookings)
            if mark_paid:
                notify_payment_confirmed(student, payment)
        except Exception:
            pass  # Never let notification failure break registration

        flash(f'Registration successful! Receipt: {payment.receipt_number} | Fee: Rs.{final}', 'success')
        return redirect(url_for('main.receipt', student_id=student.id))
    return render_template('register.html', shifts=shifts)

@main.route('/receipt/<int:student_id>')
def receipt(student_id):
    student  = Student.query.get_or_404(student_id)
    payment  = Payment.query.filter_by(student_id=student_id).first()
    bookings = Booking.query.filter_by(student_id=student_id, is_active=True).all()
    return render_template('receipt.html', student=student, payment=payment, bookings=bookings)

# ─────────────────────────────────────────────
# ADMIN AUTH
# ─────────────────────────────────────────────

@main.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if current_user.is_authenticated: return redirect(url_for('main.admin_dashboard'))
    if request.method == 'POST':
        admin = Admin.query.filter_by(username=request.form.get('username')).first()
        if admin and admin.check_password(request.form.get('password')):
            login_user(admin); return redirect(url_for('main.admin_dashboard'))
        flash('Invalid credentials.', 'danger')
    return render_template('admin_login.html')

@main.route('/admin/logout')
@login_required
def admin_logout():
    logout_user(); flash('Logged out.', 'info')
    return redirect(url_for('main.admin_login'))

# ─────────────────────────────────────────────
# ADMIN PANEL
# ─────────────────────────────────────────────

@main.route('/admin/dashboard')
@login_required
def admin_dashboard():
    due = FeeReminder.query.filter(FeeReminder.status.in_(['Pending','Overdue'])).order_by(FeeReminder.due_date).all()
    return render_template('admin_dashboard.html', stats=get_stats(), due_reminders=due)

@main.route('/admin/students')
@login_required
def admin_students():
    q        = request.args.get('q', '').strip()
    students = Student.query
    if q: students = students.filter(db.or_(Student.name.ilike(f'%{q}%'), Student.mobile.ilike(f'%{q}%')))
    students = students.order_by(Student.registered_at.desc()).all()
    return render_template('admin_students.html', students=students, q=q)

@main.route('/admin/bookings')
@login_required
def admin_bookings():
    bookings = Booking.query.filter_by(is_active=True).order_by(Booking.booked_at.desc()).all()
    return render_template('admin_bookings.html', bookings=bookings)

@main.route('/admin/booking/delete/<int:bid>', methods=['POST'])
@login_required
def delete_booking(bid):
    b = Booking.query.get_or_404(bid); b.is_active = False; db.session.commit()
    flash(f'Booking #{bid} cancelled.', 'warning')
    return redirect(url_for('main.admin_bookings'))

@main.route('/admin/student/delete/<int:sid>', methods=['POST'])
@login_required
def delete_student(sid):
    s = Student.query.get_or_404(sid); name = s.name
    db.session.delete(s); db.session.commit()
    flash(f'Student {name} removed.', 'warning')
    return redirect(url_for('main.admin_students'))

@main.route('/admin/payment/update/<int:student_id>', methods=['POST'])
def update_payment(student_id):
    p = Payment.query.filter_by(student_id=student_id).first_or_404()
    p.status = 'Paid'; p.payment_date = datetime.utcnow()
    db.session.commit()
    # Send confirmation notification
    try:
        from notifications import notify_payment_confirmed
        notify_payment_confirmed(p.student, p)
    except Exception: pass
    flash(f'Payment of Rs.{p.final_amount:.0f} marked as Paid! Receipt: {p.receipt_number}', 'success')
    ref = request.referrer or ''
    return redirect(url_for('main.receipt', student_id=student_id) if 'receipt' in ref else url_for('main.admin_students'))

@main.route('/admin/payment/edit/<int:student_id>', methods=['GET', 'POST'])
@login_required
def edit_payment(student_id):
    student = Student.query.get_or_404(student_id)
    payment = Payment.query.filter_by(student_id=student_id).first_or_404()
    if request.method == 'POST':
        dtype   = request.form.get('discount_type', 'none')
        dval    = float(request.form.get('discount_value', 0) or 0)
        dreason = request.form.get('discount_reason', '').strip()
        mode    = request.form.get('payment_mode', 'Offline')
        txn     = request.form.get('transaction_id', '').strip()
        if mode == 'Online' and not txn:
            flash('Transaction ID required.', 'danger')
            return render_template('edit_payment.html', student=student, payment=payment)
        disc, final = calc_discount(payment.original_amount, dtype, dval)
        payment.discount_type   = dtype
        payment.discount_value  = dval
        payment.discount_reason = dreason
        payment.discount_amount = disc
        payment.final_amount    = final
        payment.payment_mode    = mode
        payment.transaction_id  = txn or None
        db.session.commit()
        flash('Payment updated.', 'success')
        return redirect(url_for('main.admin_students'))
    return render_template('edit_payment.html', student=student, payment=payment)

# ─────────────────────────────────────────────
# REVENUE
# ─────────────────────────────────────────────

@main.route('/admin/revenue')
@login_required
def admin_revenue():
    payments      = Payment.query.order_by(Payment.created_at.desc()).all()
    total_paid    = db.session.query(db.func.sum(Payment.final_amount)).filter_by(status='Paid').scalar() or 0
    total_unpaid  = db.session.query(db.func.sum(Payment.final_amount)).filter_by(status='Unpaid').scalar() or 0
    total_disc    = db.session.query(db.func.sum(Payment.discount_amount)).scalar() or 0
    renewal_rev   = db.session.query(db.func.sum(FeeReminder.final_renewal_amount)).filter_by(status='Paid').scalar() or 0
    from sqlalchemy import func
    daily = db.session.query(func.date(Payment.payment_date).label('day'),
                             func.sum(Payment.final_amount).label('total')
                            ).filter(Payment.status=='Paid').group_by(func.date(Payment.payment_date)).all()
    return render_template('admin_revenue.html', payments=payments, total_paid=total_paid,
                           total_unpaid=total_unpaid, total_discount=total_disc,
                           renewal_revenue=renewal_rev, daily=daily)

# ─────────────────────────────────────────────
# FEE REMINDERS
# ─────────────────────────────────────────────

@main.route('/admin/reminders')
@login_required
def admin_reminders():
    today = date.today()
    for s in Student.query.all(): ensure_reminder(s)
    for r in FeeReminder.query.filter(FeeReminder.status=='Pending', FeeReminder.due_date < today).all():
        r.status = 'Overdue'
    db.session.commit()
    reminders = FeeReminder.query.order_by(FeeReminder.status, FeeReminder.due_date).all()
    return render_template('admin_reminders.html', reminders=reminders, today=today)

@main.route('/admin/reminders/send-all', methods=['POST'])
@login_required
def send_all_reminders():
    """Bulk send WhatsApp+SMS reminders for all pending/overdue"""
    try:
        from notifications import send_bulk_reminders
        count = send_bulk_reminders()
        flash(f'Reminders sent to {count} student(s) via WhatsApp/SMS.', 'success')
    except Exception as e:
        flash(f'Error sending reminders: {str(e)}', 'danger')
    return redirect(url_for('main.admin_reminders'))

@main.route('/admin/reminder/send/<int:rid>', methods=['POST'])
@login_required
def send_one_reminder(rid):
    """Send WhatsApp+SMS to a single student"""
    reminder = FeeReminder.query.get_or_404(rid)
    try:
        from notifications import notify_fee_reminder
        results = notify_fee_reminder(reminder.student, reminder)
        wa_ok  = results.get('whatsapp', {}).get('ok', False)
        sms_ok = results.get('sms', {}).get('ok', False)
        if wa_ok or sms_ok:
            channels = []
            if wa_ok:  channels.append('WhatsApp')
            if sms_ok: channels.append('SMS')
            flash(f"Reminder sent to {reminder.student.name} via {' & '.join(channels)}.", 'success')
        else:
            flash('Notification services not configured. Check env variables.', 'warning')
    except Exception as e:
        flash(f'Send failed: {str(e)}', 'danger')
    return redirect(url_for('main.admin_reminders'))

@main.route('/admin/reminder/collect/<int:rid>', methods=['GET', 'POST'])
@login_required
def collect_renewal(rid):
    reminder = FeeReminder.query.get_or_404(rid)
    student  = reminder.student
    if request.method == 'POST':
        disc  = float(request.form.get('discount', 0) or 0)
        mode  = request.form.get('payment_mode', 'Offline')
        txn   = request.form.get('transaction_id', '').strip()
        if mode == 'Online' and not txn:
            flash('Transaction ID required.', 'danger')
            return render_template('collect_renewal.html', reminder=reminder, student=student)
        base  = student.base_fee()
        final = max(0, base - disc)
        reminder.discount_applied       = disc
        reminder.final_renewal_amount   = final
        reminder.renewal_amount         = base
        reminder.payment_mode           = mode
        reminder.transaction_id         = txn or None
        reminder.paid_at                = datetime.utcnow()
        reminder.status                 = 'Paid'
        if student.payment:
            student.payment.next_due_date = (date.today() + relativedelta(months=1))
            student.payment.renewal_month += 1
        db.session.commit()
        # Notify student
        try:
            from notifications import notify_renewal_collected
            notify_renewal_collected(student, reminder)
        except Exception: pass
        flash(f'Renewal of Rs.{final:.0f} collected for {student.name}.', 'success')
        return redirect(url_for('main.admin_reminders'))
    return render_template('collect_renewal.html', reminder=reminder, student=student)

@main.route('/admin/reminder/waive/<int:rid>', methods=['POST'])
@login_required
def waive_reminder(rid):
    r = FeeReminder.query.get_or_404(rid); db.session.delete(r); db.session.commit()
    flash('Reminder waived.', 'info')
    return redirect(url_for('main.admin_reminders'))

# ─────────────────────────────────────────────
# NOTIFICATIONS LOG
# ─────────────────────────────────────────────

@main.route('/admin/notifications')
@login_required
def admin_notifications():
    logs = NotificationLog.query.order_by(NotificationLog.sent_at.desc()).limit(200).all()
    return render_template('admin_notifications.html', logs=logs)

@main.route('/admin/notifications/test', methods=['POST'])
@login_required
def test_notification():
    """Send a test message to verify Twilio/Fast2SMS config"""
    channel = request.form.get('channel', 'whatsapp')
    mobile  = request.form.get('mobile', '').strip()
    if len(mobile) != 10 or not mobile.isdigit():
        flash('Enter a valid 10-digit number.', 'danger')
        return redirect(url_for('main.admin_notifications'))
    msg = "SHIVA INFOTECH DIGITAL LIBRARY Test Message\nYour notification service is working correctly!"
    try:
        if channel == 'whatsapp':
            from notifications import send_via_twilio_whatsapp
            ok, info = send_via_twilio_whatsapp(mobile, msg, 0, 'test')
        elif channel == 'fast2sms':
            from notifications import send_via_fast2sms
            ok, info = send_via_fast2sms(mobile, msg, 0, 'test')
        else:
            from notifications import send_via_twilio_sms
            ok, info = send_via_twilio_sms(mobile, msg, 0, 'test')
        flash(f'{"Sent!" if ok else "Failed: " + str(info)}', 'success' if ok else 'danger')
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
    return redirect(url_for('main.admin_notifications'))

# ─────────────────────────────────────────────
# REPORTS & EXPORT
# ─────────────────────────────────────────────

@main.route('/admin/report/shift')
@login_required
def shift_report():
    shifts = Shift.query.all()
    report = []
    for s in shifts:
        bks = Booking.query.filter_by(shift_id=s.id, is_active=True).all()
        report.append({'shift': s, 'occupied': len(bks), 'available': 92-len(bks), 'bookings': bks})
    return render_template('shift_report.html', report=report)

@main.route('/admin/export/excel')
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb   = openpyxl.Workbook()
    hf   = PatternFill("solid", fgColor="1a56db")
    hfnt = Font(color="FFFFFF", bold=True)

    def hdr(ws, cols):
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = hf; cell.font = hfnt; cell.alignment = Alignment(horizontal='center')

    # Sheet 1: Students
    ws1 = wb.active; ws1.title = "Students"
    hdr(ws1, ['ID','Name','Mobile','DOB','Shifts','Seats','Base','Discount','Final','Mode','TXN','Status','Receipt','Registered'])
    for i, s in enumerate(Student.query.all(), 2):
        bks  = [b for b in s.bookings if b.is_active]
        p    = s.payment
        ws1.cell(i,1,s.id); ws1.cell(i,2,s.name); ws1.cell(i,3,s.mobile); ws1.cell(i,4,str(s.dob))
        ws1.cell(i,5,', '.join(b.shift.name for b in bks))
        ws1.cell(i,6,', '.join(str(b.seat.seat_number) for b in bks))
        ws1.cell(i,7,s.base_fee())
        ws1.cell(i,8,p.discount_display() if p else 'N/A')
        ws1.cell(i,9,p.final_amount if p else 'N/A')
        ws1.cell(i,10,p.payment_mode if p else 'N/A')
        ws1.cell(i,11,p.transaction_id or '-')
        ws1.cell(i,12,p.status if p else 'N/A')
        ws1.cell(i,13,p.receipt_number if p else 'N/A')
        ws1.cell(i,14,str(s.registered_at.date()))
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4

    # Sheet 2: Reminders
    ws2 = wb.create_sheet("Fee Reminders")
    hdr(ws2, ['ID','Student','Mobile','Month','Due','Base','Disc','Final','Mode','TXN','Status','Paid On','WA Sent','SMS Sent'])
    for i, r in enumerate(FeeReminder.query.all(), 2):
        ws2.cell(i,1,r.id); ws2.cell(i,2,r.student.name); ws2.cell(i,3,r.student.mobile)
        ws2.cell(i,4,r.month_number); ws2.cell(i,5,str(r.due_date))
        ws2.cell(i,6,r.renewal_amount); ws2.cell(i,7,r.discount_applied); ws2.cell(i,8,r.final_renewal_amount)
        ws2.cell(i,9,r.payment_mode or '-'); ws2.cell(i,10,r.transaction_id or '-')
        ws2.cell(i,11,r.status); ws2.cell(i,12,str(r.paid_at.date()) if r.paid_at else '-')
        ws2.cell(i,13,'Yes' if r.whatsapp_sent else 'No'); ws2.cell(i,14,'Yes' if r.sms_sent else 'No')

    # Sheet 3: Notification Logs
    ws3 = wb.create_sheet("Notifications")
    hdr(ws3, ['ID','Student','Mobile','Channel','Type','Status','Error','Sent At'])
    for i, n in enumerate(NotificationLog.query.all(), 2):
        ws3.cell(i,1,n.id); ws3.cell(i,2,n.student.name if n.student else '?')
        ws3.cell(i,3,n.mobile); ws3.cell(i,4,n.channel); ws3.cell(i,5,n.message_type)
        ws3.cell(i,6,n.status); ws3.cell(i,7,n.error_message or ''); ws3.cell(i,8,str(n.sent_at))

    out = io.BytesIO(); wb.save(out); out.seek(0)
    fn  = f"digilib_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(out, as_attachment=True, download_name=fn,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ─────────────────────────────────────────────
# API — health check for deployment
# ─────────────────────────────────────────────

@main.route('/health')
def health():
    return jsonify(status='ok', time=str(datetime.utcnow()))
